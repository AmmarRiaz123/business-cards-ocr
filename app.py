from flask import Flask, request, jsonify, render_template, send_file
from PIL import Image
import easyocr
import spacy
import re
import os
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from spacy.matcher import Matcher
from nameparser import HumanName

app = Flask(__name__)

# Initialize OCR and NLP tools
reader = easyocr.Reader(['en'], gpu=False)
nlp = spacy.load("en_core_web_sm")
matcher = Matcher(nlp.vocab)

# Profession keywords
profession_list = [
    "software engineer", "graphic designer", "data scientist",
    "founder", "ceo", "cto", "analyst", "consultant",
    "developer", "manager", "architect", "accountant",
    "marketing", "officer", "president", "administrator",
    "seo", "designer", "engineer"
]
profession_patterns = [[{"LOWER": token} for token in title.split()] for title in profession_list]
matcher.add("PROFESSION", profession_patterns)

# ==================== Helper Functions ====================

def preprocess_image(pil_image):
    img = np.array(pil_image)
    h, w = img.shape[:2]

    # Resize small images
    if h < 600 or w < 600:
        img = cv2.resize(img, (w * 2, h * 2), interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    adaptive = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY, 31, 10)
    kernel = np.ones((1, 1), np.uint8)
    cleaned = cv2.morphologyEx(adaptive, cv2.MORPH_OPEN, kernel)
    return cleaned

def extract_name(lines):
    for line in lines:
        name = HumanName(line.strip())
        if name.first and name.last:
            return str(name)
    return None

def extract_email(text):
    cleaned = text.replace(" ", "").replace("(@)", "@").replace("[at]", "@").replace("{at}", "@")
    cleaned = cleaned.replace("(dot)", ".").replace("[dot]", ".").replace("{dot}", ".")

    match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', cleaned)
    return match.group() if match else None

def extract_profession(doc, lines):
    matches = matcher(doc)
    if matches:
        for _, start, end in matches:
            return doc[start:end].text
    for line in lines:
        for prof in profession_list:
            if prof.lower() in line.lower():
                return line.strip()
    return None

def extract_structured_data(text):
    data = {"name": None, "email": None, "phone": None, "address": None, "profession": None}
    lines = text.split("\n")
    doc = nlp(text)

    data["name"] = extract_name(lines)
    data["email"] = extract_email(text)

    phone_match = re.search(r'(\+?\d{1,3}[\s\-]?)?(\(?\d{2,4}\)?[\s\-]?)?[\d\s\-]{7,}', text)
    if phone_match:
        data["phone"] = phone_match.group().strip()

    for ent in doc.ents:
        if ent.label_ in ["GPE", "LOC", "FAC"]:
            data["address"] = ent.text
            break

    if not data["address"]:
        for line in lines:
            if any(kw in line.lower() for kw in ["street", "road", "city", "state", "block", "avenue"]):
                data["address"] = line.strip()
                break

    data["profession"] = extract_profession(doc, lines)
    return data

def save_to_excel(data, filename="data.xlsx"):
    try:
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Email", "Phone", "Address", "Profession"])
        else:
            wb = load_workbook(filename)
            ws = wb.active

        ws.append([
            data.get("name", ""),
            data.get("email", ""),
            data.get("phone", ""),
            data.get("address", ""),
            data.get("profession", "")
        ])
        wb.save(filename)
    except Exception as e:
        print(f"[ERROR] Failed to save to Excel: {str(e)}")

# ==================== Routes ====================

@app.route('/')
def home():
    return render_template("index.html")

@app.route('/ocr', methods=['POST'])
def extract_text():
    if 'images' not in request.files:
        return jsonify({'error': 'No image files provided'}), 400

    image_files = request.files.getlist('images')
    results = []

    for image_file in image_files:
        try:
            print(f"[INFO] Processing {image_file.filename}")
            image = Image.open(image_file.stream).convert('RGB')
            preprocessed = preprocess_image(image)
            results_ocr = reader.readtext(preprocessed)

            # Retry with raw image if nothing useful is found
            if not results_ocr or all(res[2] < 0.5 for res in results_ocr):
                print(f"[WARNING] Low-confidence OCR, retrying with original image.")
                results_ocr = reader.readtext(np.array(image))

            text = '\n'.join([res[1] for res in results_ocr if res[2] > 0.4])
            if not text.strip():
                raise ValueError("OCR returned no confident text.")

            structured_data = extract_structured_data(text)
            save_to_excel(structured_data)

            results.append({
                'filename': image_file.filename,
                'raw_text': text.strip(),
                'data': structured_data
            })

        except Exception as e:
            print(f"[ERROR] Failed on {image_file.filename}: {e}")
            results.append({
                'filename': image_file.filename,
                'error': f"OCR processing failed: {str(e)}"
            })

    return jsonify(results)

@app.route('/download-excel')
def download_excel():
    path = "data.xlsx"
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    else:
        return "Excel file not found", 404

# ==================== Run Server ====================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000)
